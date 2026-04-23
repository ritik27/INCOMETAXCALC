import streamlit as st
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import os

# ---------------- SIDEBAR ----------------
st.sidebar.header("⚙️ Tax Settings")

assessment_year = st.sidebar.selectbox(
    "Select Assessment Year",
    ["AY 2026-27", "AY 2025-26", "AY 2024-25"]
)

nature = st.sidebar.selectbox(
    "Nature (Presumptive Taxation)",
    ["Normal", "44AD (Business)", "44ADA (Profession)"]
)

income_source = st.sidebar.radio(
    "Source of Income",
    ["Domestic", "Foreign"]
)

# ---------------- MAIN TITLE ----------------
st.title("💰 Income Tax Calculator (FY 2025-26)")

st.write(f"📅 Assessment Year: {assessment_year}")
st.write(f"📊 Nature: {nature}")
st.write(f"🌍 Income Source: {income_source}")

# Inputs
income = st.number_input("Enter Annual Income (₹):", min_value=0.0, step=50000.0)
deductions = st.number_input("Enter Deductions (Old Regime) (₹):", min_value=0.0, step=10000.0)

# ---------------- SPECIAL LOGIC ----------------
if nature == "44AD (Business)":
    st.info("💡 Presumptive income @ 8% applied")
    income = income * 0.08

elif nature == "44ADA (Profession)":
    st.info("💡 Presumptive income @ 50% applied")
    income = income * 0.50

if income_source == "Foreign":
    st.warning("⚠️ Foreign income may involve DTAA rules")

# ---------------- OLD REGIME ----------------
def old_tax(income, deductions):
    taxable_income = max(0, income - deductions)
    slabs = []

    if taxable_income > 300000:
        amt = min(taxable_income - 300000, 300000)
        slabs.append(("3L-6L (5%)", amt * 0.05))
    if taxable_income > 600000:
        amt = min(taxable_income - 600000, 300000)
        slabs.append(("6L-9L (10%)", amt * 0.10))
    if taxable_income > 900000:
        amt = min(taxable_income - 900000, 300000)
        slabs.append(("9L-12L (15%)", amt * 0.15))
    if taxable_income > 1200000:
        amt = min(taxable_income - 1200000, 300000)
        slabs.append(("12L-15L (20%)", amt * 0.20))
    if taxable_income > 1500000:
        amt = taxable_income - 1500000
        slabs.append(("Above 15L (30%)", amt * 0.30))

    tax = sum(x[1] for x in slabs)
    return tax, slabs

# ---------------- NEW REGIME ----------------
def new_tax(income):
    slabs = []

    if income > 400000:
        amt = min(income - 400000, 400000)
        slabs.append(("4L-8L (5%)", amt * 0.05))
    if income > 800000:
        amt = min(income - 800000, 400000)
        slabs.append(("8L-12L (10%)", amt * 0.10))
    if income > 1200000:
        amt = min(income - 1200000, 400000)
        slabs.append(("12L-16L (15%)", amt * 0.15))
    if income > 1600000:
        amt = min(income - 1600000, 400000)
        slabs.append(("16L-20L (20%)", amt * 0.20))
    if income > 2000000:
        amt = min(income - 2000000, 400000)
        slabs.append(("20L-24L (25%)", amt * 0.25))
    if income > 2400000:
        amt = income - 2400000
        slabs.append(("Above 24L (30%)", amt * 0.30))

    tax = sum(x[1] for x in slabs)
    return tax, slabs

# ---------------- CALCULATION ----------------
if st.button("Calculate Tax"):

    old, old_slabs = old_tax(income, deductions)
    new, new_slabs = new_tax(income)

    old_total = old * 1.04
    new_total = new * 1.04

    st.subheader("📊 Tax Comparison (Including 4% Cess)")

    st.write(f"🔹 Old Regime Total Tax: ₹ {old_total:,.2f}")
    st.write(f"🔹 New Regime Total Tax: ₹ {new_total:,.2f}")

    if old_total < new_total:
        st.success("✅ Old Regime is Better for You")
    elif new_total < old_total:
        st.success("✅ New Regime is Better for You")
    else:
        st.info("⚖️ Both Regimes Give Same Tax")

    # GRAPH
    st.subheader("📈 Slab-wise Tax Visualization")

    col1, col2 = st.columns(2)

    with col1:
        if old_slabs:
            fig1, ax1 = plt.subplots()
            ax1.barh([x[0] for x in old_slabs], [x[1] for x in old_slabs])
            ax1.set_title("Old Regime")
            ax1.invert_yaxis()
            st.pyplot(fig1)

    with col2:
        if new_slabs:
            fig2, ax2 = plt.subplots()
            ax2.barh([x[0] for x in new_slabs], [x[1] for x in new_slabs])
            ax2.set_title("New Regime")
            ax2.invert_yaxis()
            st.pyplot(fig2)

# ---------------- LEAD FORM (SAVE TO EXCEL) ----------------
st.markdown("---")
st.subheader("📞 Get Expert Help")

with st.form("lead_form"):
    name = st.text_input("👤 Full Name")
    phone = st.text_input("📱 Contact Number")
    email = st.text_input("📧 Email")
    message = st.text_area("📝 Query")

    submit = st.form_submit_button("📞 Request a Call")

    if submit:
        if name and phone:

            file_name = "leads.xlsx"

            # Create file if not exists
            if not os.path.exists(file_name):
                wb = Workbook()
                ws = wb.active
                ws.append(["Name", "Phone", "Email", "Query"])
                wb.save(file_name)

            # Load and append
            wb = load_workbook(file_name)
            ws = wb.active
            ws.append([name, phone, email, message])
            wb.save(file_name)

            st.success("✅ Request submitted successfully!")

        else:
            st.error("⚠️ Please fill Name and Contact Number")